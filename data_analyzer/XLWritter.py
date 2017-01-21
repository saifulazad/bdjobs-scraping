from openpyxl import Workbook

class WritterXL(object):

    def __init__(self, file_name):
        self.file_name = file_name
        self.wb = Workbook()
        self.ws = self.wb.active
        self.mapper = {'A1' : 'Company Name',
                       'B1' : 'Position',
                        'C1' : 'Hello',
                        'D1' : 'Vai'
        }
        for x in self.mapper.keys():
            self.ws[x] = self.mapper[x]


    def write_to_file(self, data):

        self.ws.append(data)
        self.wb.save(self.file_name)

    def read_from_file(self):
        return self.ws['B1'].value








# if __name__ == "__main__":
#     xl = WritterXL()
#
#     print(xl.read_from_file())
#
#     for x in range(10):
#         xl.write_to_file(['jA', 'B'])
